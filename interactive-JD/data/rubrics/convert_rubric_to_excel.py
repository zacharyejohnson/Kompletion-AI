import pandas as pd
import json
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def load_rubric(filename):
    """ Load rubric data from a JSON file """
    with open(filename, 'r') as file:
        data = json.load(file)
    return data['rubric']

def create_rubric_matrix(rubric_data):
    """ Create a pandas DataFrame from rubric data """
    categories = [item['category'] for item in rubric_data]
    weights = [item['weight'] for item in rubric_data]  # Collect weights for each category
    matrix = pd.DataFrame(index=range(1, 6), columns=categories)
    
    for category in rubric_data:
        cat_name = category['category']
        for level in category['levels']:
            score = level['score']
            description = level['description']
            matrix.at[score, cat_name] = description
    
    return matrix, weights

def save_to_excel(matrix, weights, filename):
    """ Save the DataFrame to an Excel file and format it """
    # Save DataFrame to Excel
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    matrix.to_excel(writer, index_label='Score')
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']

    # Adjust column widths and enable text wrap
    for col_num, column in enumerate(worksheet.columns, 1):
        max_len = 0
        for cell in column:
            cell_value = as_text(cell.value)  # Convert all cell values to string for processing
            len_value = len(cell_value)
            max_len = max(max_len, len_value)
        
        adjusted_width = (max_len / 3) * 1.1  # Adjust width slightly larger
        worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    

        for cell in column: 
            cell.alignment = Alignment(wrapText=True)

    worksheet.column_dimensions["A"].width = 30

        

    # Adjust row heights based on wrapped text
    for row in worksheet.iter_rows():
        max_line_count = 1
        for cell in row:
            # Convert cell value to string and count the number of lines
            cell_value = as_text(cell.value)
            line_count = len(cell_value.split('\n'))
            max_line_count = max(max_line_count, line_count)
        # Adjust the row height based on the maximum number of lines in any cell of the row
        worksheet.row_dimensions[row[0].row].height = 50 + 15 * (max_line_count)

    # Add weights row at the bottom
    last_row = len(matrix) + 2  # +1 for header, +1 for empty row
    for i, weight in enumerate(weights, 1):
        cell = worksheet.cell(row=last_row, column=i + 1)
        cell.value = f"Weight: {weight}%"

    # Add a total score row
    worksheet.cell(row=last_row + 1, column=1).value = "Total Score"
    for i in range(1, len(weights) + 1):
        col_letter = get_column_letter(i + 1)
        formula = f"=SUM({col_letter}2:{col_letter}{last_row-1}) * {weights[i-1]/100}"
        worksheet.cell(row=last_row + 1, column=i + 1).value = formula

    writer._save()

def as_text(value):
    """ Helper function to convert cell value to text """
    return str(value) if value is not None else ""

def display_and_save_rubric_matrix(json_filename, excel_filename):
    """ Load data, create matrix, print it, and save it to Excel """
    rubric_data = load_rubric(json_filename)
    matrix, weights = create_rubric_matrix(rubric_data)
    print(matrix)
    save_to_excel(matrix, weights, excel_filename)

# Assuming the JSON file is named 'rubric.json' and the Excel file will be 'rubric.xlsx'
display_and_save_rubric_matrix('interactive-JD\\data\\rubrics\\bushel_sse_rubric.json', 'interactive-JD\\data\\rubrics\\rubric.xlsx')
